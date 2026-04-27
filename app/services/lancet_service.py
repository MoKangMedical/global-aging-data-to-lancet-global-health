"""
Lancet-standard Table and Figure Generation Service.
Generates publication-quality tables (HTML/Excel) and figures (PNG).
"""
import os
import json
import numpy as np
import warnings
from typing import Dict, List, Any, Optional

warnings.filterwarnings("ignore")

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
FIGURES_DIR = os.path.join(BASE_DIR, "figures")
OUTPUTS_DIR = os.path.join(BASE_DIR, "outputs")

os.makedirs(FIGURES_DIR, exist_ok=True)
os.makedirs(OUTPUTS_DIR, exist_ok=True)

try:
    import matplotlib
    matplotlib.use("Agg")
    import matplotlib.pyplot as plt
    import matplotlib.ticker as mticker
    HAS_MATPLOTLIB = True
except ImportError:
    HAS_MATPLOTLIB = False

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


def generate_table1_html(descriptive_results: Dict[str, Any]) -> str:
    """Generate Lancet-style Table 1 (baseline characteristics) as HTML."""
    variables = descriptive_results.get("variables", {})
    n_total = descriptive_results.get("n_total", 0)
    n_exposed = descriptive_results.get("n_exposed", 0)
    n_unexposed = descriptive_results.get("n_unexposed", 0)

    display_names = {
        "age": "Age, years",
        "female": "Female sex",
        "education_low": "Low education",
        "education_mid": "Medium education",
        "education_high": "High education",
        "married": "Married/partnered",
        "smoking_current": "Current smoker",
        "smoking_former": "Former smoker",
        "diabetes": "Diabetes",
        "hypertension": "Hypertension",
        "heart_disease": "Heart disease",
        "stroke": "Stroke",
        "cancer": "Cancer",
        "depression": "Depression",
        "bmi": "Body mass index, kg/m\u00b2",
        "adl_limitation": "ADL limitation",
        "iadl_limitation": "IADL limitation",
        "cognitive_score": "Cognitive score (0-30)",
        "quality_of_life": "Quality of life (0-100)",
    }

    html = f"""
    <table class="lancet-table" style="width:100%; border-collapse:collapse; font-size:13px; font-family:Arial,sans-serif;">
        <thead>
            <tr style="background-color:#f0f0f0; border-bottom:2px solid #333;">
                <th style="text-align:left; padding:8px; border-bottom:2px solid #333;">Characteristic</th>
                <th style="text-align:center; padding:8px; border-bottom:2px solid #333;">Overall (N={n_total})</th>
                <th style="text-align:center; padding:8px; border-bottom:2px solid #333;">Exposed (N={n_exposed})</th>
                <th style="text-align:center; padding:8px; border-bottom:2px solid #333;">Unexposed (N={n_unexposed})</th>
            </tr>
        </thead>
        <tbody>
    """

    # Group continuous and binary variables
    continuous_vars = [k for k, v in variables.items() if v.get("type") == "continuous"]
    binary_vars = [k for k, v in variables.items() if v.get("type") == "binary"]

    for var in continuous_vars + binary_vars:
        if var not in variables:
            continue
        v = variables[var]
        display = display_names.get(var, var.replace("_", " ").title())
        suffix = " (SD)" if v.get("type") == "continuous" else " (n, %)"
        row_style = 'border-bottom:1px solid #ddd;'

        html += f"""
            <tr style="{row_style}">
                <td style="padding:6px 8px; text-align:left;">{display}{suffix}</td>
                <td style="padding:6px 8px; text-align:center;">{v['total']}</td>
                <td style="padding:6px 8px; text-align:center;">{v['exposed']}</td>
                <td style="padding:6px 8px; text-align:center;">{v['unexposed']}</td>
            </tr>
        """

    html += """
        </tbody>
    </table>
    <p style="font-size:11px; color:#666; margin-top:4px;">
        Data are mean (SD) or n (%). ADL=activities of daily living. IADL=instrumental activities of daily living.
    </p>
    """
    return html


def generate_regression_table_html(results: Dict[str, Any]) -> str:
    """Generate regression results table as HTML (Lancet format)."""
    method = results.get("method", "Regression")
    covariates = results.get("covariates", {})
    n_obs = results.get("n_observations", 0)
    n_events = results.get("n_events", 0)

    # Determine which columns to use
    is_survival = "hazard" in method.lower() or "cox" in method.lower() or "competing" in method.lower()
    is_logistic = "logistic" in method.lower()
    is_linear = "linear" in method.lower()

    if is_survival:
        effect_label = "HR"
        if "subdistribution" in method.lower() or "competing" in method.lower():
            effect_label = "SHR"
            lower_key = "shr_lower_95"
            upper_key = "shr_upper_95"
            effect_key = "subdistribution_hr"
        else:
            lower_key = "hr_lower_95"
            upper_key = "hr_upper_95"
            effect_key = "hazard_ratio"
    elif is_logistic:
        effect_label = "OR"
        lower_key = "or_lower_95"
        upper_key = "or_upper_95"
        effect_key = "odds_ratio"
    else:
        effect_label = "\u03b2"
        lower_key = "ci_lower_95"
        upper_key = "ci_upper_95"
        effect_key = "coef"

    html = f"""
    <table class="lancet-table" style="width:100%; border-collapse:collapse; font-size:13px; font-family:Arial,sans-serif;">
        <thead>
            <tr style="background-color:#f0f0f0; border-bottom:2px solid #333;">
                <th style="text-align:left; padding:8px; border-bottom:2px solid #333;">Variable</th>
                <th style="text-align:center; padding:8px; border-bottom:2px solid #333;">{effect_label} (95% CI)</th>
                <th style="text-align:center; padding:8px; border-bottom:2px solid #333;">P value</th>
            </tr>
        </thead>
        <tbody>
    """

    for var, vals in covariates.items():
        display = var.replace("_", " ").title()
        effect = vals.get(effect_key, vals.get("coef", "N/A"))

        if lower_key in vals and upper_key in vals:
            ci_text = f"{vals[lower_key]:.2f}–{vals[upper_key]:.2f}"
        elif "ci_lower_95" in vals:
            ci_text = f"{vals['ci_lower_95']:.2f}–{vals['ci_upper_95']:.2f}"
        else:
            ci_text = "N/A"

        p = vals.get("p_value", None)
        if p is not None:
            if p < 0.001:
                p_text = "<0\u00b7001"
            elif p < 0.01:
                p_text = f"{p:.3f}"
            else:
                p_text = f"{p:.2f}"
        else:
            p_text = "N/A"

        sig = "font-weight:bold;" if vals.get("significant", False) else ""

        html += f"""
            <tr style="border-bottom:1px solid #ddd;">
                <td style="padding:6px 8px; text-align:left; {sig}">{display}</td>
                <td style="padding:6px 8px; text-align:center; {sig}">{effect:.2f} ({ci_text})</td>
                <td style="padding:6px 8px; text-align:center; {sig}">{p_text}</td>
            </tr>
        """

    html += f"""
        </tbody>
    </table>
    <p style="font-size:11px; color:#666; margin-top:4px;">
        N={n_obs}, events={n_events}. {effect_label}=hazard ratio. CI=confidence interval.
        Bold indicates p&lt;0\u00b705.
    </p>
    """
    return html


def generate_km_curve_png(data: Dict[str, Any], project_id: int, analysis_id: Optional[int] = None) -> str:
    """Generate Kaplan-Meier survival curve as PNG. Returns file path."""
    if not HAS_MATPLOTLIB:
        return ""

    plot_data = data.get("plot_data", {})
    times = np.array(plot_data.get("time", []))
    events = np.array(plot_data.get("event", []))
    groups = plot_data.get("group", None)

    fig, ax = plt.subplots(figsize=(8, 5))
    fig.patch.set_facecolor("white")

    colors = ["#2166AC", "#B2182B", "#4DAF4A"]
    labels = ["Group 0 (Reference)", "Group 1 (Exposed)", "Group 2"]

    if groups is not None:
        groups = np.array(groups)
        for g_val, color, label in zip([0, 1], colors, labels):
            mask = groups == g_val
            if mask.sum() < 3:
                continue
            t_g = times[mask]
            e_g = events[mask]
            _plot_km_on_ax(ax, t_g, e_g, color, label)
    else:
        _plot_km_on_ax(ax, times, events, colors[0], "Overall")

    ax.set_xlabel("Time (years)", fontsize=12)
    ax.set_ylabel("Survival probability", fontsize=12)
    ax.set_title("Kaplan-Meier Survival Curves", fontsize=14, fontweight="bold")
    ax.set_ylim(0, 1.05)
    ax.set_xlim(0, max(times.max() * 1.05, 1))
    ax.legend(loc="lower left", fontsize=10)
    ax.grid(True, alpha=0.3)

    # Number at risk table below
    plt.tight_layout()

    filename = f"km_curve_p{project_id}_a{analysis_id or 0}.png"
    filepath = os.path.join(FIGURES_DIR, filename)
    fig.savefig(filepath, dpi=150, bbox_inches="tight")
    plt.close(fig)
    return filepath


def _plot_km_on_ax(ax, times, events, color, label):
    """Plot a single KM curve on given axes."""
    unique_times = np.sort(np.unique(np.append([0], np.sort(times[events == 1]))))
    n = len(times)
    surv = 1.0
    km_t = [0.0]
    km_s = [1.0]

    for t in unique_times:
        if t == 0:
            continue
        at_risk = np.sum(times >= t)
        d = np.sum((times == t) & (events == 1))
        if at_risk > 0:
            surv *= (1 - d / at_risk)
        km_t.append(float(t))
        km_s.append(float(surv))

    # Step function
    ax.step(km_t, km_s, where="post", color=color, linewidth=2, label=label)


def generate_forest_plot_png(results: Dict[str, Any], project_id: int, analysis_id: Optional[int] = None) -> str:
    """Generate forest plot of hazard ratios / odds ratios."""
    if not HAS_MATPLOTLIB:
        return ""

    covariates = results.get("covariates", {})
    if not covariates:
        return ""

    method = results.get("method", "").lower()
    if "hazard" in method or "cox" in method or "competing" in method:
        effect_key = "subdistribution_hr" if "subdistribution" in method else "hazard_ratio"
        lower_key = "shr_lower_95" if "subdistribution" in method else "hr_lower_95"
        upper_key = "shr_upper_95" if "subdistribution" in method else "hr_upper_95"
        x_label = "Hazard Ratio (95% CI)"
        ref_line = 1.0
    elif "logistic" in method:
        effect_key = "odds_ratio"
        lower_key = "or_lower_95"
        upper_key = "or_upper_95"
        x_label = "Odds Ratio (95% CI)"
        ref_line = 1.0
    else:
        effect_key = "coef"
        lower_key = "ci_lower_95"
        upper_key = "ci_upper_95"
        x_label = "Effect estimate (95% CI)"
        ref_line = 0.0

    vars_list = list(covariates.keys())
    n_vars = len(vars_list)
    fig_height = max(4, n_vars * 0.6 + 1.5)

    fig, ax = plt.subplots(figsize=(8, fig_height))
    fig.patch.set_facecolor("white")

    y_positions = list(range(n_vars - 1, -1, -1))

    for i, (var, vals) in enumerate(covariates.items()):
        effect = vals.get(effect_key, vals.get("coef", 0))
        lo = vals.get(lower_key, vals.get("ci_lower_95", 0))
        hi = vals.get(upper_key, vals.get("ci_upper_95", 0))
        p = vals.get("p_value", 1.0)
        sig = vals.get("significant", False)

        y = y_positions[i]
        color = "#2166AC" if sig else "#888888"

        # Error bar
        ax.plot([lo, hi], [y, y], color=color, linewidth=2, zorder=2)
        # Point estimate
        marker_size = 8
        ax.plot(effect, y, "D", color=color, markersize=marker_size, zorder=3)

        # Text annotation
        display = var.replace("_", " ").title()
        p_str = f"p<0.001" if p < 0.001 else f"p={p:.3f}"
        ax.text(
            ax.get_xlim()[1] if ax.get_xlim()[1] > hi else hi + 0.1,
            y,
            f"  {effect:.2f} ({lo:.2f}–{hi:.2f}), {p_str}",
            va="center", fontsize=9, color=color,
        )

    ax.set_yticks(y_positions)
    ax.set_yticklabels([v.replace("_", " ").title() for v in vars_list], fontsize=10)
    ax.axvline(ref_line, color="black", linestyle="--", linewidth=1, alpha=0.5)
    ax.set_xlabel(x_label, fontsize=11)
    ax.set_title("Forest Plot — Multivariable Analysis", fontsize=13, fontweight="bold")
    ax.spines["top"].set_visible(False)
    ax.spines["right"].set_visible(False)
    ax.grid(True, axis="x", alpha=0.2)

    plt.tight_layout()

    filename = f"forest_plot_p{project_id}_a{analysis_id or 0}.png"
    filepath = os.path.join(FIGURES_DIR, filename)
    fig.savefig(filepath, dpi=150, bbox_inches="tight")
    plt.close(fig)
    return filepath


def generate_cumulative_incidence_png(data: Dict[str, Any], project_id: int, analysis_id: Optional[int] = None) -> str:
    """Generate cumulative incidence curves for competing risks."""
    if not HAS_MATPLOTLIB:
        return ""

    plot_data = data.get("plot_data", {})
    times = np.array(plot_data.get("time", []))
    events = np.array(plot_data.get("event", []))
    competing = np.array(plot_data.get("competing_event", []))

    if len(times) == 0:
        return ""

    fig, ax = plt.subplots(figsize=(8, 5))
    fig.patch.set_facecolor("white")

    # Cumulative incidence for event of interest
    unique_times = np.sort(np.unique(times))
    n = len(times)
    ci_event = [0.0]
    ci_competing = [0.0]
    t_points = [0.0]

    cum_event = 0.0
    cum_comp = 0.0

    for t in unique_times:
        at_risk = np.sum(times >= t)
        if at_risk == 0:
            continue
        d_event = np.sum((times == t) & (events == 1))
        d_comp = np.sum((times == t) & (competing == 1))

        cum_event += d_event / n * (cum_event < 1)
        cum_comp += d_comp / n * (cum_comp < 1)

        t_points.append(float(t))
        ci_event.append(float(cum_event))
        ci_competing.append(float(cum_comp))

    ax.step(t_points, ci_event, where="post", color="#B2182B", linewidth=2,
            label="Event of interest (Cumulative incidence)")
    ax.step(t_points, ci_competing, where="post", color="#2166AC", linewidth=2,
            label="Competing event (Cumulative incidence)")

    ax.set_xlabel("Time (years)", fontsize=12)
    ax.set_ylabel("Cumulative incidence", fontsize=12)
    ax.set_title("Cumulative Incidence Functions (Competing Risks)", fontsize=14, fontweight="bold")
    ax.set_ylim(0, max(max(ci_event), max(ci_competing)) * 1.2 + 0.05)
    ax.set_xlim(0, max(times.max() * 1.05, 1))
    ax.legend(loc="upper left", fontsize=10)
    ax.grid(True, alpha=0.3)

    plt.tight_layout()

    filename = f"cumulative_incidence_p{project_id}_a{analysis_id or 0}.png"
    filepath = os.path.join(FIGURES_DIR, filename)
    fig.savefig(filepath, dpi=150, bbox_inches="tight")
    plt.close(fig)
    return filepath


def generate_tables_xlsx(project_id: int, table1_html: str, regression_html: str, analysis_results: Dict[str, Any]) -> str:
    """Generate Excel workbook with Lancet-formatted tables."""
    if not HAS_OPENPYXL:
        return ""

    wb = Workbook()

    # Styles
    header_font = Font(name="Arial", size=11, bold=True)
    normal_font = Font(name="Arial", size=10)
    header_fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
    thin_border = Border(
        bottom=Side(style="thin", color="999999"),
    )
    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")

    # Table 1 sheet
    ws1 = wb.active
    ws1.title = "Table 1"
    ws1.merge_cells("A1:D1")
    ws1["A1"] = "Table 1: Baseline Characteristics"
    ws1["A1"].font = Font(name="Arial", size=14, bold=True)

    variables = analysis_results.get("variables", {}) if "variables" in analysis_results else {}
    n_total = analysis_results.get("n_total", 0)
    n_exposed = analysis_results.get("n_exposed", 0)
    n_unexposed = analysis_results.get("n_unexposed", 0)

    headers = ["Characteristic", f"Overall (N={n_total})", f"Exposed (N={n_exposed})", f"Unexposed (N={n_unexposed})"]
    for col, h in enumerate(headers, 1):
        cell = ws1.cell(row=3, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = center_align if col > 1 else left_align

    row = 4
    for var, vals in variables.items():
        display = var.replace("_", " ").title()
        if vals.get("type") == "continuous":
            display += " (SD)"
        else:
            display += " (n, %)"

        ws1.cell(row=row, column=1, value=display).font = normal_font
        ws1.cell(row=row, column=2, value=vals["total"]).font = normal_font
        ws1.cell(row=row, column=3, value=vals["exposed"]).font = normal_font
        ws1.cell(row=row, column=4, value=vals["unexposed"]).font = normal_font
        for c in range(1, 5):
            ws1.cell(row=row, column=c).border = thin_border
            ws1.cell(row=row, column=c).alignment = center_align if c > 1 else left_align
        row += 1

    ws1.column_dimensions["A"].width = 35
    ws1.column_dimensions["B"].width = 20
    ws1.column_dimensions["C"].width = 20
    ws1.column_dimensions["D"].width = 20

    # Regression results sheet
    if analysis_results and "covariates" in analysis_results:
        ws2 = wb.create_sheet("Regression Results")
        method = analysis_results.get("method", "Regression")
        ws2.merge_cells("A1:D1")
        ws2["A1"] = f"Regression Results: {method}"
        ws2["A1"].font = Font(name="Arial", size=14, bold=True)

        reg_headers = ["Variable", "Effect (95% CI)", "P value", "Significant"]
        for col, h in enumerate(reg_headers, 1):
            cell = ws2.cell(row=3, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = center_align if col > 1 else left_align

        row = 4
        for var, vals in analysis_results.get("covariates", {}).items():
            display = var.replace("_", " ").title()
            hr = vals.get("hazard_ratio", vals.get("odds_ratio", vals.get("coef", "")))
            lo = vals.get("hr_lower_95", vals.get("or_lower_95", vals.get("ci_lower_95", "")))
            hi = vals.get("hr_upper_95", vals.get("or_upper_95", vals.get("ci_upper_95", "")))
            p = vals.get("p_value", "")
            sig = "Yes" if vals.get("significant", False) else "No"

            ws2.cell(row=row, column=1, value=display).font = normal_font
            if isinstance(hr, (int, float)):
                ws2.cell(row=row, column=2, value=f"{hr:.2f} ({lo:.2f}–{hi:.2f})").font = normal_font
            else:
                ws2.cell(row=row, column=2, value=str(hr)).font = normal_font
            ws2.cell(row=row, column=3, value=f"{p:.4f}" if isinstance(p, float) else str(p)).font = normal_font
            ws2.cell(row=row, column=4, value=sig).font = normal_font
            for c in range(1, 5):
                ws2.cell(row=row, column=c).border = thin_border
                ws2.cell(row=row, column=c).alignment = center_align if c > 1 else left_align
            row += 1

        ws2.column_dimensions["A"].width = 30
        ws2.column_dimensions["B"].width = 30
        ws2.column_dimensions["C"].width = 15
        ws2.column_dimensions["D"].width = 15

    filename = f"tables_p{project_id}.xlsx"
    filepath = os.path.join(OUTPUTS_DIR, filename)
    wb.save(filepath)
    return filepath
